"""Serviços de exportação (Excel e PDF)."""
import io
import datetime
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing
from db import fetch_all
from services.utils import format_priority_filter_label, parse_data_criacao
from config import COMPANY_NAME, DEFAULT_DEADLINE_DAYS


def get_company_name():
    """Retorna o nome da empresa para usar nos relatórios."""
    return COMPANY_NAME


def build_export_header(prioridade_filtro):
    """Constrói o cabeçalho do relatório com empresa e data."""
    company = get_company_name()
    data_geracao = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    filtro = format_priority_filter_label(prioridade_filtro)

    return {
        'company': company,
        'data_geracao': data_geracao,
        'filtro': filtro
    }


def build_demandas_data(demandas):
    """Constrói array de dados formatados para exportação."""
    data = [['ID', 'Título', 'Solicitante', 'Prioridade', 'Data Criação']]

    for demanda in demandas:
        data_partes = demanda[5].split(' ')
        date_parts = data_partes[0].split('-')
        data_formatada = f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]} {data_partes[1][:5]}"

        data.append([
            str(demanda[0]),
            demanda[1],
            demanda[3],
            demanda[4] or '---',
            data_formatada
        ])

    return data


def filter_overdue_demands(demandas, days_overdue=7):
    """Filtra demandas atrasadas."""
    now = datetime.datetime.now()
    overdue = []
    for demanda in demandas:
        created_at = parse_data_criacao(demanda[5])
        if not created_at:
            continue
        if (now - created_at).days > days_overdue:
            overdue.append(demanda)
    return overdue


def average_open_days(demandas):
    """Calcula média de dias em aberto."""
    now = datetime.datetime.now()
    deltas = []
    for demanda in demandas:
        created_at = parse_data_criacao(demanda[5])
        if not created_at:
            continue
        deltas.append((now - created_at).days)
    if not deltas:
        return 0
    return sum(deltas) / len(deltas)


def calculate_kpis(demandas):
    """Calcula KPIs para os relatórios."""
    total = len(demandas)

    criticas = sum(1 for d in demandas if d[4] and d[4].lower() in ('alta', 'crítica'))
    media = sum(1 for d in demandas if d[4] and d[4].lower() in ('média', 'media'))
    baixa = sum(1 for d in demandas if d[4] and d[4].lower() in ('baixa', 'baixo'))
    sem_prioridade = sum(1 for d in demandas if not d[4] or d[4].strip() == '')

    por_responsavel = {}
    for demanda in demandas:
        solicitante = demanda[3] or 'Sem Solicitante'
        if solicitante not in por_responsavel:
            por_responsavel[solicitante] = 0
        por_responsavel[solicitante] += 1

    por_responsavel_ordenado = sorted(por_responsavel.items(), key=lambda x: x[1], reverse=True)

    pct_criticas = (criticas / total * 100) if total > 0 else 0
    pct_media = (media / total * 100) if total > 0 else 0
    pct_baixa = (baixa / total * 100) if total > 0 else 0

    return {
        'total': total,
        'criticas': criticas,
        'pct_criticas': pct_criticas,
        'media': media,
        'pct_media': pct_media,
        'baixa': baixa,
        'pct_baixa': pct_baixa,
        'sem_prioridade': sem_prioridade,
        'por_responsavel': por_responsavel_ordenado,
    }


def fetch_demandas_for_export(prioridade_filtro='todas', relatorio_tipo='completo'):
    """Busca demandas baseado em filtros de exportação."""
    if prioridade_filtro == 'todas' or not prioridade_filtro:
        demandas = fetch_all("""
            SELECT * FROM demandas
            ORDER BY CASE LOWER(prioridade)
                         WHEN 'alta' THEN 1
                         WHEN 'média' THEN 2
                         WHEN 'media' THEN 2
                         WHEN 'baixa' THEN 3
                         ELSE 4 END,
                    data_criacao DESC
        """)
    else:
        demandas = fetch_all("""
            SELECT * FROM demandas
            WHERE LOWER(prioridade) = ?
            ORDER BY data_criacao DESC
        """, (prioridade_filtro,))

    if relatorio_tipo == 'atrasadas':
        demandas = filter_overdue_demands(demandas)

    return demandas


def export_to_excel(prioridade_filtro='todas', relatorio_tipo='completo'):
    """Exporta demandas em formato Excel."""
    demandas = fetch_demandas_for_export(prioridade_filtro, relatorio_tipo)

    header = build_export_header(prioridade_filtro)
    kpis = calculate_kpis(demandas)
    media_dias_abertas = average_open_days(demandas) if relatorio_tipo == 'atrasadas' else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Demandas"

    row = 1

    # Header
    ws[f'A{row}'] = header['company']
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = f"Gerado em: {header['data_geracao']}"
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = f"Filtro: {header['filtro']}"
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = f"Relatório: {'Demandas atrasadas' if relatorio_tipo == 'atrasadas' else 'Completo'}"
    ws.merge_cells(f'A{row}:E{row}')
    row += 2

    # KPI: Total
    ws[f'A{row}'] = "📊 TOTAL DE DEMANDAS"
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'C{row}'] = kpis['total']
    row += 1

    if media_dias_abertas is not None:
        ws[f'A{row}'] = "Média de dias em aberto (atrasadas)"
        ws[f'B{row}'] = f"{media_dias_abertas:.1f}"
        row += 1

    # KPI: Por Prioridade
    ws[f'A{row}'] = "Demandas por Prioridade"
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = "Alta (Crítica)"
    ws[f'B{row}'] = kpis['criticas']
    ws[f'C{row}'] = f"{kpis['pct_criticas']:.1f}%"
    row += 1

    ws[f'A{row}'] = "Média"
    ws[f'B{row}'] = kpis['media']
    ws[f'C{row}'] = f"{kpis['pct_media']:.1f}%"
    row += 1

    ws[f'A{row}'] = "Baixa"
    ws[f'B{row}'] = kpis['baixa']
    ws[f'C{row}'] = f"{kpis['pct_baixa']:.1f}%"
    row += 2

    # KPI: Por Responsável
    ws[f'A{row}'] = "Demandas por Responsável"
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    for responsavel, quantidade in kpis['por_responsavel']:
        ws[f'A{row}'] = responsavel
        ws[f'B{row}'] = quantidade
        row += 1

    row += 1

    # Tabela de dados
    ws[f'A{row}'] = "DETALHES DAS DEMANDAS"
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    dados = build_demandas_data(demandas)
    for row_data in dados:
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row, column=col_idx, value=cell_value)
        row += 1

    row += 1

    # Pizza: demandas por prioridade
    ws[f'A{row}'] = "DEMANDAS POR PRIORIDADE"
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    ws[f'A{row}'] = "Prioridade"
    ws[f'B{row}'] = "Quantidade"
    ws[f'C{row}'] = "Percentual"
    start_pie_row = row
    row += 1

    total = kpis['total'] or 0
    prioridade_rows = [
        ("Críticas (Alta)", kpis['criticas']),
        ("Média", kpis['media']),
        ("Baixa", kpis['baixa']),
    ]

    for label, count in prioridade_rows:
        pct = (count / total * 100) if total > 0 else 0
        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{pct:.1f}%"
        row += 1

    if total > 0:
        pie = PieChart()
        pie.title = "Demandas por Prioridade"
        data_ref = Reference(ws, min_col=2, min_row=start_pie_row + 1, max_row=start_pie_row + 3)
        labels_ref = Reference(ws, min_col=1, min_row=start_pie_row + 1, max_row=start_pie_row + 3)
        pie.add_data(data_ref, titles_from_data=False)
        pie.set_categories(labels_ref)
        pie.height = 8
        pie.width = 10
        ws.add_chart(pie, f"E{start_pie_row}")

    # Formatar larguras
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Demandas_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def export_to_pdf(prioridade_filtro='todas', relatorio_tipo='completo'):
    """Exporta demandas em formato PDF."""
    demandas = fetch_demandas_for_export(prioridade_filtro, relatorio_tipo)

    header = build_export_header(prioridade_filtro)
    kpis = calculate_kpis(demandas)
    media_dias_abertas = average_open_days(demandas) if relatorio_tipo == 'atrasadas' else None

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=36)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=6,
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#475569'),
        spaceAfter=4,
        alignment=1
    )

    kpi_title_style = ParagraphStyle(
        'KPITitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=8,
        spaceBefore=6,
    )

    story = []

    # Título
    story.append(Paragraph(header['company'], title_style))
    story.append(Spacer(1, 0.15 * inch))

    # Data e filtro
    story.append(Paragraph(f"<b>Data de geração:</b> {header['data_geracao']}", subtitle_style))
    story.append(Paragraph(f"<b>Filtro:</b> {header['filtro']}", subtitle_style))
    story.append(Paragraph(f"<b>Relatório:</b> {'Demandas atrasadas' if relatorio_tipo == 'atrasadas' else 'Completo'}", subtitle_style))
    story.append(Spacer(1, 0.25 * inch))

    # KPI 1: Total de Demandas
    story.append(Paragraph("📊 TOTAL DE DEMANDAS", kpi_title_style))
    story.append(Paragraph(f"<b style='font-size: 28'>{kpis['total']}</b>",
                          ParagraphStyle('TotalKPI', parent=styles['Normal'], alignment=1, spaceAfter=12)))
    story.append(Spacer(1, 0.1 * inch))

    if media_dias_abertas is not None:
        story.append(Paragraph(f"Média de dias em aberto (atrasadas): <b>{media_dias_abertas:.1f}</b>", subtitle_style))
        story.append(Spacer(1, 0.1 * inch))

    # KPI 2: Por Prioridade
    story.append(Paragraph("📌 DEMANDAS POR PRIORIDADE", kpi_title_style))

    priority_data = [
        ['Críticas (Alta)', f"{kpis['criticas']}", f"{kpis['pct_criticas']:.1f}%"],
        ['Média', f"{kpis['media']}", f"{kpis['pct_media']:.1f}%"],
        ['Baixa', f"{kpis['baixa']}", f"{kpis['pct_baixa']:.1f}%"],
    ]

    priority_table = Table(priority_data, colWidths=[2.5*inch, 1*inch, 1*inch])
    priority_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef3c7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#b45309')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef9f3')]),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(priority_table)
    story.append(Spacer(1, 0.2 * inch))

    # KPI 3: Demandas Críticas
    if kpis['criticas'] > 0:
        story.append(Paragraph("⚠️ DEMANDAS CRÍTICAS", ParagraphStyle(
            'CriticalTitle', parent=styles['Heading2'], fontSize=12,
            textColor=colors.HexColor('#dc2626'), spaceAfter=8, spaceBefore=6
        )))
        story.append(Paragraph(f"<b>{kpis['criticas']} demanda(s) com prioridade Alta requer(em) atenção imediata</b>",
                              subtitle_style))
        story.append(Spacer(1, 0.1 * inch))

    # KPI 4: Por Responsável
    story.append(Paragraph("👥 DEMANDAS POR RESPONSÁVEL", kpi_title_style))

    responsavel_data = [['Responsável', 'Demandas']]
    for responsavel, quantidade in kpis['por_responsavel'][:10]:
        responsavel_data.append([responsavel, str(quantidade)])

    responsavel_table = Table(responsavel_data, colWidths=[3*inch, 1*inch])
    responsavel_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(responsavel_table)
    story.append(Spacer(1, 0.25 * inch))

    # Tabela de Detalhes
    story.append(Paragraph("📋 DETALHES DAS DEMANDAS", kpi_title_style))
    dados = build_demandas_data(demandas)

    table = Table(dados, colWidths=[0.6*inch, 2.2*inch, 1.5*inch, 1.2*inch, 1.3*inch])

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbe4ff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))

    story.append(table)
    story.append(Spacer(1, 0.3 * inch))

    # Pizza
    story.append(Paragraph("DEMANDAS POR PRIORIDADE", kpi_title_style))

    priority_summary = [
        ["Críticas (Alta)", f"{kpis['criticas']}", f"{kpis['pct_criticas']:.1f}%"],
        ["Média", f"{kpis['media']}", f"{kpis['pct_media']:.1f}%"],
        ["Baixa", f"{kpis['baixa']}", f"{kpis['pct_baixa']:.1f}%"],
    ]

    priority_summary_table = Table(
        [["Prioridade", "Quantidade", "Percentual"]] + priority_summary,
        colWidths=[2.5 * inch, 1 * inch, 1 * inch],
    )
    priority_summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    story.append(priority_summary_table)

    if kpis['total'] > 0:
        drawing = Drawing(400, 200)
        drawing.hAlign = 'CENTER'
        pie = Pie()
        pie.x = 110
        pie.y = 10
        pie.width = 180
        pie.height = 180
        pie.data = [kpis['criticas'], kpis['media'], kpis['baixa']]
        pie.labels = ['Críticas', 'Média', 'Baixa']

        pie.slices.strokeWidth = 0.5
        pie.slices[0].fillColor = colors.HexColor('#ef4444')
        pie.slices[1].fillColor = colors.HexColor('#f59e0b')
        pie.slices[2].fillColor = colors.HexColor('#10b981')

        drawing.add(pie)
        story.append(Spacer(1, 0.15 * inch))
        story.append(drawing)

    story.append(Spacer(1, 0.3 * inch))

    footer_text = f"Relatório gerado com: {header['filtro']} · {datetime.datetime.now().strftime('%b–%b %Y')}"
    story.append(Paragraph(f"<i>{footer_text}</i>", subtitle_style))

    doc.build(story)
    output.seek(0)

    filename = f"Demandas_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

