import os
import sqlite3
import secrets
import datetime
from contextlib import closing
from io import BytesIO
from flask import Flask, flash, redirect, render_template, request, url_for, send_file
from flask_migrate import Migrate
from dotenv import load_dotenv
from models import db
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing

app = Flask(__name__)
load_dotenv()
flask_secret_key = os.getenv("FLASK_SECRET_KEY")

if not flask_secret_key:
    # Enforce explicit configuration in production
    if os.getenv("FLASK_ENV") == "production":
        raise RuntimeError(
            "FLASK_SECRET_KEY environment variable must be set in production."
        )
    # For non-production environments, generate a strong random key
    flask_secret_key = secrets.token_hex(32)

app.secret_key = flask_secret_key
DATABASE_PATH = os.getenv('DATABASE_PATH', 'demandas.db')
DEFAULT_DEADLINE_DAYS = 7

database_path_abs = os.path.abspath(DATABASE_PATH)
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{database_path_abs}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
migrate = Migrate(app, db)

DEFAULT_REQUESTERS = (
    {"nome": "Joao Silva", "email": "joao.silva@empresa.com", "cargo": "Analista"},
    {"nome": "Maria Santos", "email": "maria.santos@empresa.com", "cargo": "Coordenadora"},
    {"nome": "Tech Team", "email": "tech.team@empresa.com", "cargo": "Equipe Tecnica"},
    {"nome": "Equipe Suporte", "email": "suporte@empresa.com", "cargo": "Suporte"},
    {"nome": "Time Produto", "email": "produto@empresa.com", "cargo": "Produto"},
)


def get_db():
    return closing(sqlite3.connect(DATABASE_PATH))


def fetch_all(query, params=()):
    with get_db() as conn:
        cursor = conn.cursor()
        return cursor.execute(query, params).fetchall()


def fetch_one(query, params=()):
    with get_db() as conn:
        cursor = conn.cursor()
        return cursor.execute(query, params).fetchone()


def execute_query(query, params=()):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(query, params)
        conn.commit()


def log_db_error(operation, error, **context):
    context_str = ", ".join(f"{key}={value!r}" for key, value in context.items())
    app.logger.exception("Database error during %s (%s): %s", operation, context_str, error)


def date_now():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def calculate_deadline(created_at):
    created_date = parse_datetime(created_at) or datetime.datetime.now()
    return (created_date + datetime.timedelta(days=DEFAULT_DEADLINE_DAYS)).strftime("%Y-%m-%d")


def parse_datetime(value):
    if not value:
        return None
    for date_format in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(value, date_format)
        except ValueError:
            continue
    return None


def ensure_demandas_dashboard_columns():
    columns = {row[1] for row in fetch_all("PRAGMA table_info(demandas)")}
    required_columns = {
        "status": "TEXT NOT NULL DEFAULT 'aberta'",
        "responsavel": "TEXT",
        "prazo": "TEXT",
        "data_conclusao": "TEXT",
    }

    for column_name, column_definition in required_columns.items():
        if column_name not in columns:
            execute_query(f"ALTER TABLE demandas ADD COLUMN {column_name} {column_definition}")

    execute_query(
        """
        UPDATE demandas
        SET responsavel = solicitante
        WHERE responsavel IS NULL OR TRIM(responsavel) = ''
        """
    )
    execute_query(
        f"""
        UPDATE demandas
        SET prazo = COALESCE(
            date(data_criacao, '+{DEFAULT_DEADLINE_DAYS} days'),
            date('now', '+{DEFAULT_DEADLINE_DAYS} days')
        )
        """
    )
    execute_query(
        """
        UPDATE demandas
        SET prioridade = 'baixa'
        WHERE prioridade IS NULL
           OR TRIM(prioridade) = ''
           OR LOWER(TRIM(prioridade)) NOT IN ('alta', 'media', 'média', 'baixa')
        """
    )


def normalize_status(value):
    status = (value or "aberta").strip().lower()
    if status in ("concluida", "concluída"):
        return "concluida"
    if status == "cancelada":
        return "cancelada"
    return "aberta"


def normalize_priority(value):
    priority = (value or "").strip().lower()
    if priority == "alta":
        return "alta"
    if priority in ("media", "média"):
        return "media"
    return "baixa"


def status_label(status):
    labels = {
        "aberta": "Aberta",
        "concluida": "Concluida",
        "cancelada": "Cancelada",
    }
    return labels.get(normalize_status(status), "Aberta")


def calculate_dashboard_metrics(demandas):
    total = len(demandas)
    today = datetime.datetime.now().date()

    counts = {
        "abertas": 0,
        "concluidas": 0,
        "atrasadas": 0,
    }
    critical_demands = []
    open_by_responsible = {}
    completed_resolution_days = []

    for demanda in demandas:
        status = normalize_status(demanda[6] if len(demanda) > 6 else None)
        responsavel = (demanda[7] if len(demanda) > 7 else None) or demanda[3] or "Sem responsavel"
        prazo = demanda[8] if len(demanda) > 8 else None
        data_conclusao = demanda[9] if len(demanda) > 9 else None

        is_finalized = status in ("concluida", "cancelada")
        is_late = False
        prazo_date = parse_datetime(prazo)
        if prazo_date and not is_finalized and prazo_date.date() < today:
            is_late = True

        if status == "concluida":
            counts["concluidas"] += 1
            started_at = parse_datetime(demanda[5])
            finished_at = parse_datetime(data_conclusao)
            if started_at and finished_at and finished_at >= started_at:
                completed_resolution_days.append((finished_at - started_at).total_seconds() / 86400)
        elif status == "aberta":
            counts["abertas"] += 1
            open_by_responsible[responsavel] = open_by_responsible.get(responsavel, 0) + 1

        if is_late:
            counts["atrasadas"] += 1

        if (demanda[4] or "").strip().lower() == "alta" and not is_finalized:
            critical_demands.append(
                {
                    "id": demanda[0],
                    "titulo": demanda[1],
                    "responsavel": responsavel,
                    "status": status_label(status),
                }
            )

    def percent(value):
        if total == 0:
            return 0
        return round((value / total) * 100)

    average_resolution_days = None
    if completed_resolution_days:
        average_resolution_days = round(sum(completed_resolution_days) / len(completed_resolution_days), 1)

    return {
        "total": total,
        "status_cards": [
            {"label": "Abertas", "count": counts["abertas"], "percent": percent(counts["abertas"])},
            {"label": "Concluidas", "count": counts["concluidas"], "percent": percent(counts["concluidas"])},
            {"label": "Atrasadas", "count": counts["atrasadas"], "percent": percent(counts["atrasadas"])},
        ],
        "critical_demands": critical_demands,
        "open_by_responsible": sorted(open_by_responsible.items(), key=lambda item: (-item[1], item[0].lower())),
        "average_resolution_days": average_resolution_days,
    }


def ensure_requesters_table_seeded():
    execute_query(
        """
        CREATE TABLE IF NOT EXISTS requesters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            cargo TEXT NOT NULL,
            data_criacao TEXT NOT NULL
        )
        """
    )

    row = fetch_one("SELECT COUNT(1) FROM requesters")
    count = row[0] if row else 0

    if count == 0:
        now = date_now()
        for requester in DEFAULT_REQUESTERS:
            execute_query(
                """
                INSERT INTO requesters (nome, email, cargo, data_criacao)
                VALUES (?, ?, ?, ?)
                """,
                (requester["nome"], requester["email"], requester["cargo"], now),
            )


def get_requesters():
    ensure_requesters_table_seeded()
    rows = fetch_all(
        """
        SELECT id, nome, email, cargo
        FROM requesters
        ORDER BY nome COLLATE NOCASE ASC
        """
    )
    return [{"id": row[0], "nome": row[1], "email": row[2], "cargo": row[3]} for row in rows]


def get_requester_by_id(requester_id):
    ensure_requesters_table_seeded()
    row = fetch_one(
        """
        SELECT id, nome, email, cargo, data_criacao
        FROM requesters
        WHERE id = ?
        """,
        (requester_id,),
    )
    if not row:
        return None
    return {"id": row[0], "nome": row[1], "email": row[2], "cargo": row[3], "data_criacao": row[4]}


def get_available_people():
    # Placeholder for the future login integration that will hydrate this list.
    normalized_people = []

    for requester in get_requesters():
        name = str(requester.get("nome", "")).strip()
        if name and name not in normalized_people:
            normalized_people.append(name)

    return normalized_people


def is_valid_person(name):
    return name in get_available_people()


@app.route('/', methods=['GET'])
def index():
    ensure_demandas_dashboard_columns()
    prioridade_filtro = request.args.get('prioridade', 'todas').strip().lower()
    if prioridade_filtro not in ("todas", ""):
        prioridade_filtro = normalize_priority(prioridade_filtro)
    
    if prioridade_filtro == 'todas' or not prioridade_filtro:
        demandas = fetch_all("""
                             SELECT *
                             FROM demandas
                             ORDER BY CASE LOWER(prioridade)
                                          WHEN 'alta' THEN 1
                                          WHEN 'média' THEN 2
                                          WHEN 'media' THEN 2
                                          WHEN 'baixa' THEN 3
                                          ELSE 4
                                          END,
                                      data_criacao DESC
                             """)
    else:
        demandas = fetch_all("""
                             SELECT *
                             FROM demandas
                             WHERE LOWER(prioridade) = ?
                             ORDER BY data_criacao DESC
                             """, (prioridade_filtro,))
    
    return render_template('index.html', demandas=demandas, prioridade_filtro=prioridade_filtro)


@app.route('/dashboard', methods=['GET'])
def dashboard():
    ensure_demandas_dashboard_columns()
    dashboard_metrics = calculate_dashboard_metrics(fetch_all("SELECT * FROM demandas"))
    return render_template('dashboard.html', dashboard=dashboard_metrics)

@app.route('/nova_demanda', methods=['GET', 'POST'])
def nova_demanda():
    ensure_demandas_dashboard_columns()
    people = get_available_people()

    if request.method == 'POST':
        titulo = request.form['titulo']
        descricao = request.form['descricao']
        solicitante = request.form['solicitante'].strip()
        responsavel = request.form.get('responsavel', '').strip()
        prioridade = normalize_priority(request.form.get('prioridade'))
        status = normalize_status(request.form.get('status'))
        data_criacao = str(date_now())
        prazo = calculate_deadline(data_criacao)
        data_conclusao = date_now() if status == "concluida" else None

        if not is_valid_person(solicitante) or not is_valid_person(responsavel):
            flash('Selecione solicitante e responsavel validos!', 'error')
            return render_template(
                'nova_demanda.html',
                people=people,
                form_data=request.form,
            )

        try:
            execute_query(
                """
                INSERT INTO demandas
                    (titulo, descricao, solicitante, prioridade, data_criacao, status, responsavel, prazo, data_conclusao)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (titulo, descricao, solicitante, prioridade, data_criacao, status, responsavel, prazo, data_conclusao),
            )
        except sqlite3.Error as error:
            log_db_error("nova_demanda", error, solicitante=solicitante)
            flash('Erro ao salvar demanda!')
            return redirect(url_for('nova_demanda'))

        flash('Salvo!')
        return redirect(url_for('index'))

    return render_template('nova_demanda.html', people=people, form_data={})


@app.route('/editar/<int:demanda_id>', methods=['GET', 'POST'])
def editar(demanda_id):
    ensure_demandas_dashboard_columns()
    people = get_available_people()

    if request.method == 'POST':
        titulo = request.form['titulo']
        descricao = request.form['descricao']
        prioridade = normalize_priority(request.form.get('prioridade'))
        solicitante = request.form['solicitante'].strip()
        responsavel = request.form.get('responsavel', '').strip()
        status = normalize_status(request.form.get('status'))
        current_demanda = fetch_one('SELECT * FROM demandas WHERE id = ?', (demanda_id,))
        prazo = calculate_deadline(current_demanda[5] if current_demanda else None)
        current_data_conclusao = current_demanda[9] if current_demanda and len(current_demanda) > 9 else None
        data_conclusao = current_data_conclusao
        if status == "concluida" and not data_conclusao:
            data_conclusao = date_now()
        elif status != "concluida":
            data_conclusao = None

        if not is_valid_person(solicitante) or not is_valid_person(responsavel):
            flash('Selecione solicitante e responsavel validos!', 'error')
            demanda = fetch_one('SELECT * FROM demandas WHERE id = ?', (demanda_id,))
            return render_template(
                'editar.html',
                demanda=demanda,
                people=people,
                form_data=request.form,
            )

        try:
            execute_query(
                """
                UPDATE demandas
                SET titulo = ?, descricao = ?, prioridade = ?, solicitante = ?,
                    responsavel = ?, status = ?, prazo = ?, data_conclusao = ?
                WHERE id = ?
                """,
                (titulo, descricao, prioridade, solicitante, responsavel, status, prazo, data_conclusao, demanda_id),
            )
        except sqlite3.Error as error:
            log_db_error("editar", error, demanda_id=demanda_id, solicitante=solicitante)
            flash('Erro ao atualizar demanda!')
            return redirect(url_for('editar', demanda_id=demanda_id))

        flash('Atualizado!')
        return redirect(url_for('index'))

    demanda = fetch_one('SELECT * FROM demandas WHERE id = ?', (demanda_id,))
    return render_template('editar.html', demanda=demanda, people=people, form_data={})


@app.route('/deletar/<int:demanda_id>', methods=['DELETE'])
def deletar(demanda_id):
    try:
        execute_query('DELETE FROM demandas WHERE id = ?', (demanda_id,))
    except sqlite3.Error as error:
        log_db_error("deletar", error, demanda_id=demanda_id)
        flash('Erro ao deletar demanda!')
        return redirect(url_for('index'))

    flash('Deletado!')
    return redirect(url_for('index'))


@app.route('/buscar', methods=['GET'])
def buscar():
    ensure_demandas_dashboard_columns()
    termo = request.args.get('q', '').strip()
    like_term = f"%{termo}%"

    resultados = fetch_all(
        '''SELECT * FROM demandas 
           WHERE titulo LIKE ? 
           OR id LIKE ?
           OR solicitante LIKE ?''',
        (like_term, like_term, like_term),
    )
    return render_template('index.html', demandas=resultados, prioridade_filtro='todas')


@app.route('/detalhes/<int:demanda_id>', methods=['GET'])
def detalhes(demanda_id):
    ensure_demandas_dashboard_columns()
    demanda = fetch_one('SELECT * FROM demandas WHERE id = ?', (demanda_id,))
    comentarios = fetch_all('SELECT * FROM comentarios WHERE demanda_id = ?', (demanda_id,))

    return render_template(
        'detalhes.html',
        demanda=demanda,
        comentarios=comentarios,
        people=get_available_people(),
        form_data={},
    )


@app.route('/solicitante', methods=['GET'])
def solicitante():
    requesters = get_requesters()
    
    # Count de chamados por solicitante
    dados = fetch_all("""
        SELECT solicitante, COUNT(*) as total
        FROM demandas
        GROUP BY solicitante
        ORDER BY total DESC
    """)
    
    # Formata para o gráfico
    categorias = [row[0] for row in dados]
    totais = [row[1] for row in dados]
    
    return render_template('solicitante.html', 
                         requesters=requesters,
                         categorias=categorias,
                         totais=totais)


@app.route('/sutita', methods=['GET'])
def sutita():
    return render_template('sutita.html')


@app.route('/solicitante/<int:requester_id>', methods=['GET'])
def detalhes_solicitante(requester_id):
    requester = get_requester_by_id(requester_id)
    if not requester:
        flash('Solicitante não encontrado!', 'error')
        return redirect(url_for('solicitante'))
    return render_template('detalhes_solicitante.html', requester=requester)


@app.route('/solicitante/editar/<int:requester_id>', methods=['GET', 'POST'])
def editar_solicitante(requester_id):
    requester = get_requester_by_id(requester_id)
    if not requester:
        flash('Solicitante não encontrado!', 'error')
        return redirect(url_for('solicitante'))

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        cargo = request.form.get('cargo', '').strip()

        if not nome or not email or not cargo:
            flash('Preencha nome, email e cargo para salvar!', 'error')
            return render_template('editar_solicitante.html', requester=requester, form_data=request.form)

        if len(nome) > 40 or len(email) > 40 or len(cargo) > 40:
            flash('Nome, email e cargo devem ter no máximo 40 caracteres.', 'error')
            return render_template('editar_solicitante.html', requester=requester, form_data=request.form)

        try:
            execute_query(
                """
                UPDATE requesters
                SET nome = ?, email = ?, cargo = ?
                WHERE id = ?
                """,
                (nome, email, cargo, requester_id),
            )
        except sqlite3.IntegrityError:
            flash('Já existe solicitante cadastrado com este email!', 'error')
            return render_template('editar_solicitante.html', requester=requester, form_data=request.form)
        except sqlite3.Error as error:
            log_db_error("editar_solicitante", error, requester_id=requester_id, email=email)
            flash('Erro ao editar solicitante!', 'error')
            return render_template('editar_solicitante.html', requester=requester, form_data=request.form)

        flash('Solicitante atualizado com sucesso!', 'success')
        return redirect(url_for('solicitante'))

    return render_template('editar_solicitante.html', requester=requester, form_data={})


@app.route('/solicitante/deletar/<int:requester_id>', methods=['DELETE'])
def deletar_solicitante(requester_id):
    requester = get_requester_by_id(requester_id)
    if not requester:
        return ('', 404)

    try:
        execute_query("DELETE FROM requesters WHERE id = ?", (requester_id,))
    except sqlite3.Error as error:
        log_db_error("deletar_solicitante", error, requester_id=requester_id)
        return ('', 500)

    return ('', 204)


@app.route('/solicitante/cadastrar', methods=['GET', 'POST'])
def cadastrar_solicitante():
    if request.method == 'POST':
        nome = request.form['nome'].strip()
        email = request.form['email'].strip()
        cargo = request.form['cargo'].strip()

        if not nome or not email or not cargo:
            flash('Preencha nome, email e cargo para cadastrar!', 'error')
            return render_template('cadastrar_solicitante.html', form_data=request.form)

        if len(nome) > 40 or len(email) > 40 or len(cargo) > 40:
            flash('Nome, email e cargo devem ter no máximo 40 caracteres.', 'error')
            return render_template('cadastrar_solicitante.html', form_data=request.form)

        try:
            execute_query(
                """
                INSERT INTO requesters (nome, email, cargo, data_criacao)
                VALUES (?, ?, ?, ?)
                """,
                (nome, email, cargo, str(date_now())),
            )
        except sqlite3.IntegrityError:
            flash('Ja existe solicitante cadastrado com este email!', 'error')
            return render_template('cadastrar_solicitante.html', form_data=request.form)
        except sqlite3.Error as error:
            log_db_error("cadastrar_solicitante", error, nome=nome, email=email)
            flash('Erro ao cadastrar solicitante!', 'error')
            return render_template('cadastrar_solicitante.html', form_data=request.form)

        flash('Solicitante cadastrado com sucesso!', 'success')
        return redirect(url_for('solicitante'))

    return render_template('cadastrar_solicitante.html', form_data={})


@app.route('/adicionar_comentario/<int:demanda_id>', methods=['POST'])
def adicionar_comentario(demanda_id):
    comentario = request.form['comentario'].strip()
    autor = request.form['autor'].strip()

    if not is_valid_person(autor):
        flash('Selecione um autor valido para o comentario!', 'error')
        demanda = fetch_one('SELECT * FROM demandas WHERE id = ?', (demanda_id,))
        comentarios = fetch_all('SELECT * FROM comentarios WHERE demanda_id = ?', (demanda_id,))
        return render_template(
            'detalhes.html',
            demanda=demanda,
            comentarios=comentarios,
            people=get_available_people(),
            form_data=request.form,
        )

    if not comentario:
        flash('Escreva um comentario antes de enviar!', 'error')
        demanda = fetch_one('SELECT * FROM demandas WHERE id = ?', (demanda_id,))
        comentarios = fetch_all('SELECT * FROM comentarios WHERE demanda_id = ?', (demanda_id,))
        return render_template(
            'detalhes.html',
            demanda=demanda,
            comentarios=comentarios,
            people=get_available_people(),
            form_data=request.form,
        )

    try:
        execute_query(
            'INSERT INTO comentarios (demanda_id, comentario, autor, data) VALUES (?, ?, ?, ?)',
            (demanda_id, comentario, autor, str(date_now())),
        )
    except sqlite3.Error as error:
        log_db_error("adicionar_comentario", error, demanda_id=demanda_id, autor=autor)
        flash('Erro ao adicionar comentário!')

    return redirect(url_for('detalhes', demanda_id=demanda_id))


def get_company_name():
    """Retorna o nome da empresa para usar nos relatórios"""
    return os.getenv("COMPANY_NAME", "SGDI - Sistema de Gestão de Demandas")


def format_priority_filter_label(prioridade_filtro):
    """Formata o rótulo do filtro de prioridade para exibição"""
    if prioridade_filtro == 'todas' or not prioridade_filtro:
        return "Todas as prioridades"
    elif prioridade_filtro == 'alta':
        return "Prioridade Alta"
    elif prioridade_filtro in ('media', 'média'):
        return "Prioridade Média"
    elif prioridade_filtro == 'baixa':
        return "Prioridade Baixa"
    return prioridade_filtro


def build_export_header(prioridade_filtro):
    """Constrói o cabeçalho do relatório com empresa e data"""
    company = get_company_name()
    data_geracao = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    filtro = format_priority_filter_label(prioridade_filtro)
    
    return {
        'company': company,
        'data_geracao': data_geracao,
        'filtro': filtro
    }


def build_demandas_data(demandas):
    """Constrói array de dados formatados para exportação"""
    data = [['ID', 'Título', 'Solicitante', 'Prioridade', 'Data Criação']]

    for demanda in demandas:
        data_partes = demanda[5].split(' ')
        date_parts = data_partes[0].split('-')
        data_formatada = f"{date_parts[2]}/{date_parts[1]}/{date_parts[0]} {data_partes[1][:5]}"

        data.append([
            str(demanda[0]),
            demanda[1],
            demanda[3],
            demanda[4] or '---',
            data_formatada
        ])

    return data


def parse_data_criacao(value):
    if not value:
        return None
    try:
        return datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def filter_overdue_demands(demandas, days_overdue=7):
    now = datetime.datetime.now()
    overdue = []
    for demanda in demandas:
        created_at = parse_data_criacao(demanda[5])
        if not created_at:
            continue
        if (now - created_at).days > days_overdue:
            overdue.append(demanda)
    return overdue


def average_open_days(demandas):
    now = datetime.datetime.now()
    deltas = []
    for demanda in demandas:
        created_at = parse_data_criacao(demanda[5])
        if not created_at:
            continue
        deltas.append((now - created_at).days)
    if not deltas:
        return 0
    return sum(deltas) / len(deltas)


def calculate_kpis(demandas):
    """Calcula KPIs para os relatórios"""
    total = len(demandas)

    # Contar por prioridade (assumindo: Alta=crítica, Média=normal, Baixa=baixa prioridade)
    criticas = sum(1 for d in demandas if d[4] and d[4].lower() in ('alta', 'crítica'))
    media = sum(1 for d in demandas if d[4] and d[4].lower() in ('média', 'media'))
    baixa = sum(1 for d in demandas if d[4] and d[4].lower() in ('baixa', 'baixo'))
    sem_prioridade = sum(1 for d in demandas if not d[4] or d[4].strip() == '')

    # Contar por responsável
    por_responsavel = {}
    for demanda in demandas:
        solicitante = demanda[3] or 'Sem Solicitante'
        if solicitante not in por_responsavel:
            por_responsavel[solicitante] = 0
        por_responsavel[solicitante] += 1

    # Ordenar por quantidade (descendente)
    por_responsavel_ordenado = sorted(por_responsavel.items(), key=lambda x: x[1], reverse=True)

    # Calcular percentuais
    pct_criticas = (criticas / total * 100) if total > 0 else 0
    pct_media = (media / total * 100) if total > 0 else 0
    pct_baixa = (baixa / total * 100) if total > 0 else 0

    return {
        'total': total,
        'criticas': criticas,
        'pct_criticas': pct_criticas,
        'media': media,
        'pct_media': pct_media,
        'baixa': baixa,
        'pct_baixa': pct_baixa,
        'sem_prioridade': sem_prioridade,
        'por_responsavel': por_responsavel_ordenado,
    }


@app.route('/export/excel', methods=['POST'])
def export_excel():
    """Exporta demandas em formato Excel"""
    prioridade_filtro = request.form.get('prioridade_filtro', 'todas').strip().lower()
    relatorio_tipo = request.form.get('relatorio_tipo', 'completo').strip().lower()

    # Fetch dados
    if prioridade_filtro == 'todas' or not prioridade_filtro:
        demandas = fetch_all("""
            SELECT * FROM demandas
            ORDER BY CASE LOWER(prioridade)
                         WHEN 'alta' THEN 1
                         WHEN 'média' THEN 2
                         WHEN 'media' THEN 2
                         WHEN 'baixa' THEN 3
                         ELSE 4 END,
                    data_criacao DESC
        """)
    else:
        demandas = fetch_all("""
            SELECT * FROM demandas
            WHERE LOWER(prioridade) = ?
            ORDER BY data_criacao DESC
        """, (prioridade_filtro,))

    if relatorio_tipo == 'atrasadas':
        demandas = filter_overdue_demands(demandas)

    # Montar header e KPIs
    header = build_export_header(prioridade_filtro)
    kpis = calculate_kpis(demandas)
    media_dias_abertas = average_open_days(demandas) if relatorio_tipo == 'atrasadas' else None

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Demandas"

    row = 1

    # Header com informações da empresa
    ws[f'A{row}'] = header['company']
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = f"Gerado em: {header['data_geracao']}"
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = f"Filtro: {header['filtro']}"
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = f"Relatório: {'Demandas atrasadas' if relatorio_tipo == 'atrasadas' else 'Completo'}"
    ws.merge_cells(f'A{row}:E{row}')
    row += 2

    # KPI 1: Total de demandas
    ws[f'A{row}'] = "📊 TOTAL DE DEMANDAS"
    ws[f'A{row}'].font = ws[f'A{row}'].font.copy()
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'C{row}'] = kpis['total']
    row += 1

    if media_dias_abertas is not None:
        ws[f'A{row}'] = "Média de dias em aberto (atrasadas)"
        ws[f'B{row}'] = f"{media_dias_abertas:.1f}"
        row += 1

    # KPI 2: Por Prioridade
    ws[f'A{row}'] = "Demandas por Prioridade"
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    ws[f'A{row}'] = "Alta (Crítica)"
    ws[f'B{row}'] = kpis['criticas']
    ws[f'C{row}'] = f"{kpis['pct_criticas']:.1f}%"
    row += 1

    ws[f'A{row}'] = "Média"
    ws[f'B{row}'] = kpis['media']
    ws[f'C{row}'] = f"{kpis['pct_media']:.1f}%"
    row += 1

    ws[f'A{row}'] = "Baixa"
    ws[f'B{row}'] = kpis['baixa']
    ws[f'C{row}'] = f"{kpis['pct_baixa']:.1f}%"
    row += 2

    # KPI 3: Por Responsável
    ws[f'A{row}'] = "Demandas por Responsável"
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    for responsavel, quantidade in kpis['por_responsavel']:
        ws[f'A{row}'] = responsavel
        ws[f'B{row}'] = quantidade
        row += 1

    row += 1

    # Tabela de dados
    ws[f'A{row}'] = "DETALHES DAS DEMANDAS"
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    dados = build_demandas_data(demandas)
    for row_data in dados:
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row, column=col_idx, value=cell_value)
        row += 1

    row += 1

    # Pizza: demandas por prioridade (percentual)
    ws[f'A{row}'] = "DEMANDAS POR PRIORIDADE"
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    ws[f'A{row}'] = "Prioridade"
    ws[f'B{row}'] = "Quantidade"
    ws[f'C{row}'] = "Percentual"
    start_pie_row = row
    row += 1

    total = kpis['total'] or 0
    prioridade_rows = [
        ("Críticas (Alta)", kpis['criticas']),
        ("Média", kpis['media']),
        ("Baixa", kpis['baixa']),
    ]

    for label, count in prioridade_rows:
        pct = (count / total * 100) if total > 0 else 0
        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{pct:.1f}%"
        row += 1

    if total > 0:
        pie = PieChart()
        pie.title = "Demandas por Prioridade"
        data_ref = Reference(ws, min_col=2, min_row=start_pie_row + 1, max_row=start_pie_row + 3)
        labels_ref = Reference(ws, min_col=1, min_row=start_pie_row + 1, max_row=start_pie_row + 3)
        pie.add_data(data_ref, titles_from_data=False)
        pie.set_categories(labels_ref)
        pie.height = 8
        pie.width = 10
        ws.add_chart(pie, f"E{start_pie_row}")

    # Formatar larguras
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    # Salvar em bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Demandas_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/export/pdf', methods=['POST'])
def export_pdf():
    """Exporta demandas em formato PDF"""
    prioridade_filtro = request.form.get('prioridade_filtro', 'todas').strip().lower()
    relatorio_tipo = request.form.get('relatorio_tipo', 'completo').strip().lower()

    # Fetch dados
    if prioridade_filtro == 'todas' or not prioridade_filtro:
        demandas = fetch_all("""
            SELECT * FROM demandas
            ORDER BY CASE LOWER(prioridade)
                         WHEN 'alta' THEN 1
                         WHEN 'média' THEN 2
                         WHEN 'media' THEN 2
                         WHEN 'baixa' THEN 3
                         ELSE 4 END,
                    data_criacao DESC
        """)
    else:
        demandas = fetch_all("""
            SELECT * FROM demandas
            WHERE LOWER(prioridade) = ?
            ORDER BY data_criacao DESC
        """, (prioridade_filtro,))

    if relatorio_tipo == 'atrasadas':
        demandas = filter_overdue_demands(demandas)

    # Montar header e KPIs
    header = build_export_header(prioridade_filtro)
    kpis = calculate_kpis(demandas)
    media_dias_abertas = average_open_days(demandas) if relatorio_tipo == 'atrasadas' else None

    # Criar PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=36)

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=6,
        alignment=1  # Center
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#475569'),
        spaceAfter=4,
        alignment=1  # Center
    )

    kpi_title_style = ParagraphStyle(
        'KPITitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=8,
        spaceBefore=6,
    )

    kpi_value_style = ParagraphStyle(
        'KPIValue',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=4,
        fontName='Helvetica-Bold',
    )

    # Conteúdo
    story = []

    # Título
    story.append(Paragraph(header['company'], title_style))
    story.append(Spacer(1, 0.15 * inch))

    # Data e filtro
    story.append(Paragraph(f"<b>Data de geração:</b> {header['data_geracao']}", subtitle_style))
    story.append(Paragraph(f"<b>Filtro:</b> {header['filtro']}", subtitle_style))
    story.append(Paragraph(f"<b>Relatório:</b> {'Demandas atrasadas' if relatorio_tipo == 'atrasadas' else 'Completo'}", subtitle_style))
    story.append(Spacer(1, 0.25 * inch))

    # KPI 1: Total de Demandas (destaque)
    story.append(Paragraph("📊 TOTAL DE DEMANDAS", kpi_title_style))
    story.append(Paragraph(f"<b style='font-size: 28'>{kpis['total']}</b>", 
                          ParagraphStyle('TotalKPI', parent=styles['Normal'], alignment=1, spaceAfter=12)))
    story.append(Spacer(1, 0.1 * inch))

    if media_dias_abertas is not None:
        story.append(Paragraph(f"Média de dias em aberto (atrasadas): <b>{media_dias_abertas:.1f}</b>", subtitle_style))
        story.append(Spacer(1, 0.1 * inch))

    # KPI 2: Por Prioridade
    story.append(Paragraph("📌 DEMANDAS POR PRIORIDADE", kpi_title_style))

    priority_data = [
        ['Críticas (Alta)', f"{kpis['criticas']}", f"{kpis['pct_criticas']:.1f}%"],
        ['Média', f"{kpis['media']}", f"{kpis['pct_media']:.1f}%"],
        ['Baixa', f"{kpis['baixa']}", f"{kpis['pct_baixa']:.1f}%"],
    ]

    priority_table = Table(priority_data, colWidths=[2.5*inch, 1*inch, 1*inch])
    priority_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef3c7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#b45309')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef9f3')]),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(priority_table)
    story.append(Spacer(1, 0.2 * inch))

    # KPI 3: Demandas Críticas (Atenção)
    if kpis['criticas'] > 0:
        story.append(Paragraph("⚠️ DEMANDAS CRÍTICAS", ParagraphStyle(
            'CriticalTitle', parent=styles['Heading2'], fontSize=12, 
            textColor=colors.HexColor('#dc2626'), spaceAfter=8, spaceBefore=6
        )))

        story.append(Paragraph(f"<b>{kpis['criticas']} demanda(s) com prioridade Alta requer(em) atenção imediata</b>",
                              subtitle_style))
        story.append(Spacer(1, 0.1 * inch))

    # KPI 4: Por Responsável
    story.append(Paragraph("👥 DEMANDAS POR RESPONSÁVEL", kpi_title_style))

    responsavel_data = [['Responsável', 'Demandas']]
    for responsavel, quantidade in kpis['por_responsavel'][:10]:  # Limitar a 10 primeiros
        responsavel_data.append([responsavel, str(quantidade)])

    responsavel_table = Table(responsavel_data, colWidths=[3*inch, 1*inch])
    responsavel_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(responsavel_table)
    story.append(Spacer(1, 0.25 * inch))

    # Tabela de Detalhes
    story.append(Paragraph("📋 DETALHES DAS DEMANDAS", kpi_title_style))
    dados = build_demandas_data(demandas)

    table = Table(dados, colWidths=[0.6*inch, 2.2*inch, 1.5*inch, 1.2*inch, 1.3*inch])

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbe4ff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))

    story.append(table)
    story.append(Spacer(1, 0.3 * inch))

    # Pizza: demandas por prioridade (percentual)
    story.append(Paragraph("DEMANDAS POR PRIORIDADE", kpi_title_style))

    priority_summary = [
        ["Críticas (Alta)", f"{kpis['criticas']}", f"{kpis['pct_criticas']:.1f}%"],
        ["Média", f"{kpis['media']}", f"{kpis['pct_media']:.1f}%"],
        ["Baixa", f"{kpis['baixa']}", f"{kpis['pct_baixa']:.1f}%"],
    ]

    priority_summary_table = Table(
        [["Prioridade", "Quantidade", "Percentual"]] + priority_summary,
        colWidths=[2.5 * inch, 1 * inch, 1 * inch],
    )
    priority_summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    story.append(priority_summary_table)

    if kpis['total'] > 0:
        drawing = Drawing(400, 200)
        drawing.hAlign = 'CENTER'
        pie = Pie()
        pie.x = 110
        pie.y = 10
        pie.width = 180
        pie.height = 180
        pie.data = [kpis['criticas'], kpis['media'], kpis['baixa']]
        pie.labels = ['Críticas', 'Média', 'Baixa']
        
        # Melhorando o visual (cor e bordas)
        pie.slices.strokeWidth = 0.5
        pie.slices[0].fillColor = colors.HexColor('#ef4444') # Vermelho para Críticas
        pie.slices[1].fillColor = colors.HexColor('#f59e0b') # Laranja para Média
        pie.slices[2].fillColor = colors.HexColor('#10b981') # Verde para Baixa
        
        drawing.add(pie)
        story.append(Spacer(1, 0.15 * inch))
        story.append(drawing)

    story.append(Spacer(1, 0.3 * inch))

    # Rodapé com filtro aplicado
    footer_text = f"Relatório gerado com: {header['filtro']} · {datetime.datetime.now().strftime('%b–%b %Y')}"
    story.append(Paragraph(f"<i>{footer_text}</i>", subtitle_style))

    # Build PDF
    doc.build(story)
    output.seek(0)

    filename = f"Demandas_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':
    host = os.getenv("FLASK_HOST", "127.0.0.1")
    port = int(os.getenv("FLASK_PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "false").strip().lower() in ("1", "true", "yes", "on")
    app.run(debug=debug, host=host, port=port)
